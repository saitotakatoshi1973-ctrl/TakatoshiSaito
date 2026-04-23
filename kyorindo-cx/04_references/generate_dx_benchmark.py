import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ---- カラー定義 ----
C_HEADER_DS   = "1F3864"
C_HEADER_GMS  = "1A5276"
C_HEADER_REF  = "4A235A"
C_COL_TITLE   = "2E4057"
C_ROW_DS      = "EBF5FB"
C_ROW_GMS     = "EAF4F4"
C_ROW_REF     = "F5EEF8"
C_STAR5       = "27AE60"
C_STAR4       = "82E0AA"
C_STAR3       = "F9E79F"
C_STAR2       = "F0B27A"
C_STAR1       = "EC7063"
C_WHITE       = "FFFFFF"
C_MIZEN       = "FADBD8"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

STAR_COLOR = {
    "★★★★★": C_STAR5,
    "★★★★":  C_STAR4,
    "★★★":   C_STAR3,
    "★★":    C_STAR2,
    "★":     C_STAR1,
}

def row_fill_color(cat):
    if "DS" in cat:     return C_ROW_DS
    elif "参考" in cat: return C_ROW_REF
    else:               return C_ROW_GMS

# ---- トピックデータ ----
# (企業名, カテゴリ, 静岡県内規模, DX成熟度, 分類, 施策・取り組み名, 実装時期, 主な導入目的, 成果・実績, ツルハHD比較)
TOPICS = [
    # ウエルシア薬局
    ("ウエルシア薬局", "直接競合（DS）", "約215店", "★★★★★", "AI活用", "ダイナミックプライシング（564商品）", "2023年頃", "粗利改善・在庫最適化", "対象商品粗利+30%以上", "未着手"),
    ("ウエルシア薬局", "直接競合（DS）", "約215店", "★★★★★", "AI活用", "AI売上予測「SalesSensor」", "2025年度", "需要予測精度向上・販売計画最適化", "2025年度導入", "未着手"),
    ("ウエルシア薬局", "直接競合（DS）", "約215店", "★★★★★", "AI活用", "カスタマーAI（NEL）", "2024年", "顧客対応工数削減・CS品質向上", "工数90%削減", "未着手"),
    ("ウエルシア薬局", "直接競合（DS）", "約215店", "★★★★★", "データ基盤", "Treasure Data CDP", "2022年頃", "医療・健康・POSデータ統合・MAツール実装", "データ統合完了・MAツール実装済み", "未着手"),
    ("ウエルシア薬局", "直接競合（DS）", "約215店", "★★★★★", "アプリ・デジタル", "混雑チェックアプリ", "不明", "来店体験向上・混雑回避", "実装済み", "不明"),
    ("ウエルシア薬局", "直接競合（DS）", "約215店", "★★★★★", "アプリ・デジタル", "モバイルTカード", "不明", "デジタル会員証化・レジ待ち削減", "実装済み", "同等"),
    ("ウエルシア薬局", "直接競合（DS）", "約215店", "★★★★★", "アプリ・デジタル", "1to1マーケティング（True Data）", "不明", "購買行動に基づくパーソナライズ提案", "実装済み", "未着手"),
    ("ウエルシア薬局", "直接競合（DS）", "約215店", "★★★★★", "調剤DX", "オンライン服薬指導", "不明", "患者利便性向上・遠隔医療対応", "実装済み", "不明"),
    ("ウエルシア薬局", "直接競合（DS）", "約215店", "★★★★★", "リテールメディア", "デジタルサイネージ", "2023年（1,600店）", "店内広告収益化・購買促進", "1,600店導入済み", "不明"),
    ("ウエルシア薬局", "直接競合（DS）", "約215店", "★★★★★", "EX（従業員体験）", "業務端末リプレイス（2026年）", "2026年予定", "現場業務効率化・デジタル化推進", "2026年完了予定", "不明"),
    # クリエイトエス・ディー
    ("クリエイトエス・ディー", "直接競合（DS）", "約95〜100店", "★★★", "アプリ・デジタル", "公式アプリ（ポイントカード・クーポン・チラシ）", "不明", "顧客接点デジタル化・来店促進", "実装済み", "同等"),
    ("クリエイトエス・ディー", "直接競合（DS）", "約95〜100店", "★★★", "調剤DX", "処方箋事前送信（アプリ経由）", "不明", "調剤待ち時間削減・患者利便性向上", "実装済み", "不明"),
    # スギ薬局
    ("スギ薬局", "直接競合（DS）", "約35〜37店", "★★★★★", "AI活用", "品揃え最適化AI（エクサウィザーズ共同開発）", "不明", "在庫最適化・品揃え精度向上", "実装済み（詳細非公開）", "未着手"),
    ("スギ薬局", "直接競合（DS）", "約35〜37店", "★★★★★", "AI活用", "薬剤師アシスタントAI", "2026年4月（本番化）", "薬剤師の問い合わせ対応効率化", "200店舗展開・2026年4月本番化", "未着手"),
    ("スギ薬局", "直接競合（DS）", "約35〜37店", "★★★★★", "データ基盤", "データレイクハウス（AWS）", "構築中", "社内分散データ統合・データドリブン経営", "構築中", "未着手"),
    ("スギ薬局", "直接競合（DS）", "約35〜37店", "★★★★★", "アプリ・デジタル", "スギ薬局アプリ", "不明", "既存顧客の深化・LTV向上", "実装済み", "同等"),
    ("スギ薬局", "直接競合（DS）", "約35〜37店", "★★★★★", "アプリ・デジタル", "アバター遠隔接客", "2022年頃", "問い合わせ対応効率化・省人化", "対応時間1/4に短縮", "未着手"),
    ("スギ薬局", "直接競合（DS）", "約35〜37店", "★★★★★", "アプリ・デジタル", "生活習慣病リスクレポート", "不明", "ヘルスケア提案強化・健康管理サービス", "実装済み", "未着手"),
    ("スギ薬局", "直接競合（DS）", "約35〜37店", "★★★★★", "調剤DX", "スギスマホでお薬", "不明", "処方箋デジタル化・待ち時間削減", "実装済み", "不明"),
    ("スギ薬局", "直接競合（DS）", "約35〜37店", "★★★★★", "リテールメディア", "リテールメディア「お茶の間」戦略", "不明", "購買データ活用でメーカー向け広告収益化", "購買率前年比144%", "未着手"),
    ("スギ薬局", "直接競合（DS）", "約35〜37店", "★★★★★", "EX（従業員体験）", "生成AI QAボット（年末調整）", "2024年頃", "問い合わせ対応工数削減・ナレッジ共有", "1ヶ月で構築・200店展開", "同等"),
    # マツキヨコクミン
    ("マツキヨコクミン", "直接競合（DS）", "約30〜32店", "★★★★★", "AI活用", "商品DNA（意識スコアタグ付け）", "不明", "購買データ活用・PB商品開発精度向上", "新規顧客9割のPB開発実現", "未着手"),
    ("マツキヨコクミン", "直接競合（DS）", "約30〜32店", "★★★★★", "データ基盤", "リアル×デジタル統合プラットフォーム", "不明", "顧客データ統合・接点最大化", "会員2,342万人・接点1億5,000万超", "未着手"),
    ("マツキヨコクミン", "直接競合（DS）", "約30〜32店", "★★★★★", "アプリ・デジタル", "ARメイクシミュレーター「Be Makeup+」", "不明", "購買前のデジタル体験提供・EC誘導", "実装済み", "未着手"),
    ("マツキヨコクミン", "直接競合（DS）", "約30〜32店", "★★★★★", "アプリ・デジタル", "肌診断「Be Skincare+」・髪診断「Be Haircare+」", "不明", "スキンケア・ヘアケア提案強化", "実装済み", "未着手"),
    ("マツキヨコクミン", "直接競合（DS）", "約30〜32店", "★★★★★", "アプリ・デジタル", "店舗発送型EC「マツキヨコクミンQ」", "不明", "即日配送で利便性向上・EC収益拡大", "21都道府県展開", "未着手"),
    ("マツキヨコクミン", "直接競合（DS）", "約30〜32店", "★★★★★", "調剤DX", "処方箋事前送信・電子お薬手帳（マツキヨコクミンMe）", "不明", "調剤業務デジタル化・患者利便性向上", "実装済み", "不明"),
    ("マツキヨコクミン", "直接競合（DS）", "約30〜32店", "★★★★★", "リテールメディア", "Matsukiyo Ads（メーカー向け広告プラットフォーム）", "不明", "ファーストパーティデータ活用での広告収益化", "購買データ×Google広告で運用", "未着手"),
    # クスリのアオキ
    ("クスリのアオキ", "直接競合（DS）", "静岡進出中（拡大中）", "★★★", "AI活用", "AIチャットボット「WisTalk」", "不明", "社内問い合わせ対応自動化・業務効率化", "業務負荷75%削減", "不明"),
    ("クスリのアオキ", "直接競合（DS）", "静岡進出中（拡大中）", "★★★", "データ基盤", "クラウドPOS（TMN）", "不明", "販売データのリアルタイム活用・分析", "実装済み", "不明"),
    ("クスリのアオキ", "直接競合（DS）", "静岡進出中（拡大中）", "★★★", "データ基盤", "SuperStream-NX Cloud", "不明", "月次決算早期化・財務DXの推進", "実装済み", "不明"),
    ("クスリのアオキ", "直接競合（DS）", "静岡進出中（拡大中）", "★★★", "アプリ・デジタル", "公式アプリ（Aocaポイント・処方箋送信・電子お薬手帳）", "不明", "顧客接点デジタル化・利便性向上", "実装済み", "同等"),
    ("クスリのアオキ", "直接競合（DS）", "静岡進出中（拡大中）", "★★★", "調剤DX", "処方箋スマホ送信・電子お薬手帳（アプリ内）", "不明", "調剤待ち時間削減・ペーパーレス化", "実装済み", "不明"),
    # ゲンキー
    ("ゲンキー", "直接競合（DS）", "東海エリア展開中", "★★", "その他", "（DX情報確認できていない）", "不明", "低価格食品訴求型。DXより価格競争力で差別化", "確認できていない", "不明"),
    # コスモス薬品
    ("コスモス薬品", "直接競合（DS）", "進出未確認", "★★★", "データ基盤", "POS×棚割り連動システム", "不明", "陳列最適化・フェース数管理の徹底", "全店一元管理実装済み", "不明"),
    ("コスモス薬品", "直接競合（DS）", "進出未確認", "★★★", "データ基盤", "本部一元管理型自動発注システム", "不明", "在庫最適化・人的作業削減", "全店展開済み", "不明"),
    ("コスモス薬品", "直接競合（DS）", "進出未確認", "★★★", "アプリ・デジタル", "アプリ内オンラインストア（アプリ限定商品・クーポン）", "不明", "顧客接点拡大・EC収益化", "実装済み", "同等"),
    ("コスモス薬品", "直接競合（DS）", "進出未確認", "★★★", "EX（従業員体験）", "本部×店舗の役割分担明確化", "不明", "オペレーション効率化・店舗負荷軽減", "全社展開済み", "不明"),
    # サンドラッグ
    ("サンドラッグ", "直接競合（DS）", "進出未確認", "★★★★", "AI活用", "AI FAQシステム（ユーザーローカル）", "不明", "カスタマーサポート効率化・問い合わせ自動応答", "実装済み", "不明"),
    ("サンドラッグ", "直接競合（DS）", "進出未確認", "★★★★", "AI活用", "AIカメラ実証実験（Ultimatrust）", "不明", "来客動態分析・店舗改善", "実証実験中", "不明"),
    ("サンドラッグ", "直接競合（DS）", "進出未確認", "★★★★", "データ基盤", "Salesforce Service Cloud＋Experience Cloud", "不明", "EC・店舗コミュニケーション一元管理", "実装済み", "不明"),
    ("サンドラッグ", "直接競合（DS）", "進出未確認", "★★★★", "アプリ・デジタル", "Shopify Plus ECサイトリニューアル", "不明", "EC体験向上・売上拡大", "実装済み", "不明"),
    ("サンドラッグ", "直接競合（DS）", "進出未確認", "★★★★", "リテールメディア", "Criteoリテールメディア", "不明", "ファーストパーティデータ活用でメーカー向け広告収益化", "業界先駆け・先行優位を確立", "未着手"),
    ("サンドラッグ", "直接競合（DS）", "進出未確認", "★★★★", "リテールメディア", "トクスルビジョン（デジタルサイネージ）", "2025年3月（全1,000店）", "店内広告のデジタル化・広告収益化", "全1,000店・2025年3月導入完了", "不明"),
    # カワチ薬品
    ("カワチ薬品", "直接競合（DS）", "進出未確認", "★★", "その他", "（DX情報確認できていない）", "不明", "大型郊外店型。静岡への進出は現時点で未確認", "確認できていない", "不明"),
    # マックスバリュ東海
    ("マックスバリュ東海（イオン）", "間接競合（GMS）", "約77店（県内GMS1位）", "★★★★", "AI活用", "AIワーク・MaIボード", "不明", "店舗業務効率化・AI活用推進", "約350店導入", "不明"),
    ("マックスバリュ東海（イオン）", "間接競合（GMS）", "約77店（県内GMS1位）", "★★★★", "AI活用", "生成AI推進（グループ1,000人トライアル）", "2023年頃", "生成AIスキル育成・業務効率化", "グループ全体でトライアル中", "同等"),
    ("マックスバリュ東海（イオン）", "間接競合（GMS）", "約77店（県内GMS1位）", "★★★★", "データ基盤", "イオングループデータ統合基盤", "継続中", "グループ横断でのデータ活用・デジタル売上拡大", "デジタル売上高1兆円目標（2026年度）", "不明"),
    ("マックスバリュ東海（イオン）", "間接競合（GMS）", "約77店（県内GMS1位）", "★★★★", "アプリ・デジタル", "WAONアプリ等グループ共通基盤への統合", "2025年2月", "ポイント統合・顧客接点の一元化", "独自アプリ終了→グループ統合完了", "不明"),
    ("マックスバリュ東海（イオン）", "間接競合（GMS）", "約77店（県内GMS1位）", "★★★★", "調剤DX", "ウエルシアとのグループデータ連携", "検討中", "食品×調剤のクロスユース促進", "検討・推進中", "不明"),
    ("マックスバリュ東海（イオン）", "間接競合（GMS）", "約77店（県内GMS1位）", "★★★★", "リテールメディア", "イオングループリテールメディア展開", "継続中", "グループデータ活用での広告収益化", "グループ全体で展開中", "不明"),
    # イオン
    ("イオン（GMS旗艦）", "間接競合（GMS）", "約7店", "★★★★", "AI活用", "AIワーク・MaIボード", "不明", "店舗業務効率化・AI活用推進", "約350店導入", "不明"),
    ("イオン（GMS旗艦）", "間接競合（GMS）", "約7店", "★★★★", "AI活用", "生成AI内製化推進", "継続中", "内製DXスキル育成・コスト削減", "グループ全体で推進中", "不明"),
    ("イオン（GMS旗艦）", "間接競合（GMS）", "約7店", "★★★★", "データ基盤", "グループ顧客データ統合基盤", "継続中", "デジタル売上高拡大・顧客データ活用", "1兆円目標（2026年度）", "不明"),
    ("イオン（GMS旗艦）", "間接競合（GMS）", "約7店", "★★★★", "アプリ・デジタル", "iAEONアプリ（WAON統合）", "不明", "決済・ポイント統合・顧客接点拡大", "グループ共通アプリとして展開", "不明"),
    ("イオン（GMS旗艦）", "間接競合（GMS）", "約7店", "★★★★", "アプリ・デジタル", "ネットスーパー強化", "継続中", "EC・即配送で利便性向上・顧客獲得", "拡大中", "不明"),
    ("イオン（GMS旗艦）", "間接競合（GMS）", "約7店", "★★★★", "リテールメディア", "イオンメディアグループリテールメディア", "継続中", "グループデータ活用での広告収益化", "グループ展開中", "不明"),
    # アピタ
    ("アピタ（ユニー／PPIH）", "間接競合（GMS）", "静岡・東海展開", "★★★", "データ基盤", "PPIHグループ共通デジタル基盤", "不明", "グループデータ活用・DX推進", "グループ展開中", "不明"),
    ("アピタ（ユニー／PPIH）", "間接競合（GMS）", "静岡・東海展開", "★★★", "アプリ・デジタル", "majicaアプリ（PPIH共通ポイント）", "不明", "ポイント統合・顧客接点拡大", "グループ共通アプリとして展開", "不明"),
    # バロー
    ("バロー（バローHD）", "間接競合（GMS）", "東海エリア展開", "★★★", "データ基盤", "グループ内データ共有クラウドシステム", "不明", "商品・販売データのリアルタイム共有", "約1,300店舗に展開済み", "不明"),
    ("バロー（バローHD）", "間接競合（GMS）", "東海エリア展開", "★★★", "データ基盤", "製造小売業DX（3つのコネクト戦略）", "継続中", "製造〜流通〜販売の一元管理", "推進中", "不明"),
    # 遠鉄ストア
    ("遠鉄ストア", "間接競合（GMS）", "浜松市中心に展開", "★★★", "データ基盤", "RETAILSTUDIO®（売場・販促管理）", "2023年1月", "売場・販促管理のデジタル化・効率化", "2023年1月導入", "不明"),
    ("遠鉄ストア", "間接競合（GMS）", "浜松市中心に展開", "★★★", "アプリ・デジタル", "LINEミニアプリ（デジタル会員証）", "2023年2月", "デジタル会員証化・来店促進", "2023年2月導入（静岡県西部スーパー初）", "未着手"),
    ("遠鉄ストア", "間接競合（GMS）", "浜松市中心に展開", "★★★", "アプリ・デジタル", "スマートレシート", "不明", "ペーパーレスレシート・購買履歴のデジタル管理", "実装済み", "不明"),
    # 田子重
    ("田子重", "間接競合（スーパー）", "焼津市本部・静岡県内", "★", "その他", "（DX情報確認できていない）", "不明", "地元密着・生鮮強みでの顧客囲い込みに注力", "確認できていない", "不明"),
    # ドン・キホーテ
    ("ドン・キホーテ（PPIH）", "間接競合（ディスカウント）", "県内複数店", "★★★★", "AI活用", "AIプライシング（価格最適化）", "不明", "競争力強化・粗利管理の自動化", "実装済み（詳細非公開）", "未着手"),
    ("ドン・キホーテ（PPIH）", "間接競合（ディスカウント）", "県内複数店", "★★★★", "データ基盤", "内製DXチームによる業務自動化基盤", "継続中", "現場業務の自動化・効率化", "年間300万時間削減", "未着手"),
    ("ドン・キホーテ（PPIH）", "間接競合（ディスカウント）", "県内複数店", "★★★★", "アプリ・デジタル", "majicaアプリ（PPIH共通ポイント）", "不明", "ポイント統合・顧客接点拡大", "グループ共通アプリとして展開", "不明"),
    ("ドン・キホーテ（PPIH）", "間接競合（ディスカウント）", "県内複数店", "★★★★", "アプリ・デジタル", "無人小型店舗「キャンパスドンキ」", "実証中", "無人化・新業態の実証", "AIカメラ×重量センサーで実証中", "未着手"),
    # 業務スーパー
    ("業務スーパー（神戸物産）", "間接競合（ディスカウント）", "県内展開", "★", "その他", "（DX情報確認できていない）", "不明", "調達・製造コスト削減で価格競争力を確保", "確認できていない", "不明"),
    # ツルハHD
    ("ツルハHD（参考：親会社）", "参考（グループ親会社）", "杏林堂経由で全店", "★★★", "AI活用", "生成AIナレッジ検索（NECと共同開発）", "2026年2月（1,000店完了）", "従業員の情報アクセス効率化・ナレッジ共有", "1,000店舗展開完了（2026年2月）", "-"),
    ("ツルハHD（参考：親会社）", "参考（グループ親会社）", "杏林堂経由で全店", "★★★", "データ基盤", "業界初データクリーンルーム「ツルハDCR」", "2025年2月", "リテールメディア強化・ファーストパーティデータ活用", "2025年2月構築", "-"),
    ("ツルハHD（参考：親会社）", "参考（グループ親会社）", "杏林堂経由で全店", "★★★", "データ基盤", "SmartDB（契約管理・内部統制）", "不明", "契約管理のデジタル化・内部統制強化", "実装済み", "-"),
    ("ツルハHD（参考：親会社）", "参考（グループ親会社）", "杏林堂経由で全店", "★★★", "アプリ・デジタル", "スマホアプリ（楽天ポイント連携・個別商品提案）", "不明", "顧客接点デジタル化・購買促進", "実装済み", "-"),
    ("ツルハHD（参考：親会社）", "参考（グループ親会社）", "杏林堂経由で全店", "★★★", "リテールメディア", "データクリーンルーム活用リテールメディア", "2025年〜継続中", "広告収益化・購買データ提供", "リテールAIアワード2024受賞", "-"),
    ("ツルハHD（参考：親会社）", "参考（グループ親会社）", "杏林堂経由で全店", "★★★", "EX（従業員体験）", "生成AIナレッジ検索（画像・図表対応）", "2026年2月（1,000店完了）", "従業員の業務効率化・ナレッジ活用", "1,000店舗展開完了（2026年2月）", "-"),
]

# ---- シート1: 競合DXベンチマーク（トピック行形式） ----
ws1 = wb.active
ws1.title = "競合DXベンチマーク"

COLUMNS = [
    "企業名", "カテゴリ", "静岡県内規模", "DX成熟度",
    "分類", "施策・取り組み名", "実装時期", "主な導入目的", "成果・実績", "ツルハHD比較"
]
COL_WIDTHS = [22, 22, 22, 12, 18, 32, 18, 34, 30, 12]

# タイトル行
ws1.merge_cells("A1:J1")
c = ws1["A1"]
c.value = "杏林堂 競合DXベンチマーク一覧（2026年4月23日）　1行＝1施策・取り組み"
c.font = Font(bold=True, color="FFFFFF", size=13)
c.fill = fill("1A3A5C")
c.alignment = align("center", "center")
ws1.row_dimensions[1].height = 30

# 列ヘッダー行
for ci, col_name in enumerate(COLUMNS, 1):
    c = ws1.cell(row=2, column=ci, value=col_name)
    c.font = Font(bold=True, color="FFFFFF", size=10)
    c.fill = fill(C_COL_TITLE)
    c.alignment = align("center", "center")
    c.border = thin_border()
ws1.row_dimensions[2].height = 30

# 列幅
for i, w in enumerate(COL_WIDTHS, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# データ書き込み（企業ごとにセル結合）
current_row = 3

company_groups = {}
company_order = []
for t in TOPICS:
    name = t[0]
    if name not in company_groups:
        company_groups[name] = []
        company_order.append(name)
    company_groups[name].append(t)

for company_name in company_order:
    topics = company_groups[company_name]
    start_row = current_row
    n = len(topics)

    for i, t in enumerate(topics):
        (name, cat, shizuoka, star, classification, measure, timing, purpose, result, tsurha) = t
        row_color = row_fill_color(cat)

        # 企業名（後でマージ）
        c_name = ws1.cell(row=current_row, column=1)
        if i == 0:
            c_name.value = name
            c_name.font = Font(bold=True, size=10)
        c_name.fill = fill(row_color)
        c_name.border = thin_border()
        c_name.alignment = align("center", "center")

        # カテゴリ（後でマージ）
        c_cat = ws1.cell(row=current_row, column=2)
        if i == 0:
            c_cat.value = cat
            c_cat.font = Font(size=9)
        c_cat.fill = fill(row_color)
        c_cat.border = thin_border()
        c_cat.alignment = align("center", "center")

        # 静岡県内規模（後でマージ）
        c_shi = ws1.cell(row=current_row, column=3)
        if i == 0:
            c_shi.value = shizuoka
            c_shi.font = Font(size=9)
        c_shi.fill = fill(row_color)
        c_shi.border = thin_border()
        c_shi.alignment = align("center", "center")

        # DX成熟度（後でマージ）
        c_star = ws1.cell(row=current_row, column=4)
        if i == 0:
            c_star.value = star
            c_star.font = Font(bold=True, size=11,
                               color="FFFFFF" if star in ["★★★★★", "★★★★"] else "333333")
        c_star.fill = fill(STAR_COLOR.get(star, C_STAR1))
        c_star.border = thin_border()
        c_star.alignment = align("center", "center")

        # 分類
        c_cls = ws1.cell(row=current_row, column=5, value=classification)
        c_cls.font = Font(bold=True, size=9)
        c_cls.fill = fill(row_color)
        c_cls.border = thin_border()
        c_cls.alignment = align("center", "center")

        # 施策・取り組み名
        c_msr = ws1.cell(row=current_row, column=6, value=measure)
        c_msr.font = Font(size=9)
        c_msr.fill = fill(row_color)
        c_msr.border = thin_border()
        c_msr.alignment = align("left", "center")

        # 実装時期
        c_tim = ws1.cell(row=current_row, column=7, value=timing)
        c_tim.font = Font(size=9)
        c_tim.fill = fill(row_color)
        c_tim.border = thin_border()
        c_tim.alignment = align("center", "center")

        # 主な導入目的
        c_pur = ws1.cell(row=current_row, column=8, value=purpose)
        c_pur.font = Font(size=9)
        c_pur.fill = fill(row_color)
        c_pur.border = thin_border()
        c_pur.alignment = align("left", "center")

        # 成果・実績
        c_res = ws1.cell(row=current_row, column=9, value=result)
        c_res.font = Font(size=9)
        c_res.fill = fill(row_color)
        c_res.border = thin_border()
        c_res.alignment = align("left", "center")

        # ツルハHD比較
        c_tsr = ws1.cell(row=current_row, column=10, value=tsurha)
        c_tsr.font = Font(bold=True, size=9,
                          color="CC0000" if tsurha == "未着手" else "333333")
        c_tsr.fill = fill(C_MIZEN if tsurha == "未着手" else row_color)
        c_tsr.border = thin_border()
        c_tsr.alignment = align("center", "center")

        ws1.row_dimensions[current_row].height = 45
        current_row += 1

    # 企業名・カテゴリ・静岡・DX成熟度のセル結合
    if n > 1:
        end_row = start_row + n - 1
        for col in [1, 2, 3, 4]:
            ws1.merge_cells(
                start_row=start_row, start_column=col,
                end_row=end_row, end_column=col
            )

# ---- シート2: DX成熟度サマリー ----
ws2 = wb.create_sheet("DX成熟度サマリー")

SUMMARY = [
    ("ウエルシア薬局", "直接競合（DS）", "約215店", "★★★★★", "データ統合×AI販促×カスタマーAI（工数90%削減）"),
    ("スギ薬局", "直接競合（DS）", "約35〜37店", "★★★★★", "AIファースト宣言・薬剤師AI・リテールメディア（購買率+144%）"),
    ("マツキヨコクミン", "直接競合（DS）", "約30〜32店", "★★★★★", "商品DNA・ARビューティー・Matsukiyo Ads（リテールメディア事業化最先進）"),
    ("サンドラッグ", "直接競合（DS）", "進出未確認", "★★★★", "リテールメディア先駆け（Criteo）・トクスルビジョン1,000店"),
    ("ドン・キホーテ（PPIH）", "間接競合（ディスカウント）", "県内複数店", "★★★★", "AIプライシング・内製DX・年間300万時間削減"),
    ("マックスバリュ東海（イオン）", "間接競合（GMS）", "約77店（県内GMS1位）", "★★★★", "グループデータ統合・デジタル売上1兆円目標・ウエルシア連携"),
    ("イオン（GMS旗艦）", "間接競合（GMS）", "約7店", "★★★★", "グループデータ統合・イオンスタイル静岡2026年春開業"),
    ("ツルハHD（参考：親会社）", "参考（グループ親会社）", "杏林堂経由で全店", "★★★", "生成AIナレッジ検索（1,000店完了）・データクリーンルーム（業界初）"),
    ("クリエイトエス・ディー", "直接競合（DS）", "約95〜100店", "★★★", "アプリ・処方箋送信（標準的）"),
    ("クスリのアオキ", "直接競合（DS）", "静岡進出中（拡大中）", "★★★", "AIチャットボット（負荷75%削減）・クラウドPOS・アプリ"),
    ("コスモス薬品", "直接競合（DS）", "進出未確認", "★★★", "POS×棚割り・自動発注（オペレーション効率特化）"),
    ("バロー（バローHD）", "間接競合（GMS）", "東海エリア展開", "★★★", "グループデータ共有基盤・製造小売DX（1,300店展開）"),
    ("アピタ（ユニー／PPIH）", "間接競合（GMS）", "静岡・東海展開", "★★★", "PPIHグループ施策（majicaアプリ）"),
    ("遠鉄ストア", "間接競合（GMS）", "浜松市中心に展開", "★★★", "LINEミニアプリ（県西部スーパー初）・RETAILSTUDIO®"),
    ("ゲンキー", "直接競合（DS）", "東海エリア展開中", "★★", "情報少なく評価困難（低価格食品訴求型）"),
    ("カワチ薬品", "直接競合（DS）", "進出未確認", "★★", "情報少なく評価困難（大型郊外店型）"),
    ("田子重", "間接競合（スーパー）", "焼津市本部・静岡県内", "★", "DX情報なし（地元密着・生鮮強み）"),
    ("業務スーパー（神戸物産）", "間接競合（ディスカウント）", "県内展開", "★", "DX情報なし（調達コスト削減特化）"),
]

SUM_COLS = ["順位", "企業名", "カテゴリ", "静岡県内規模", "DX成熟度", "最大の強み"]
SUM_WIDTHS = [8, 28, 24, 24, 14, 56]

ws2.merge_cells("A1:F1")
c = ws2["A1"]
c.value = "DX成熟度サマリー（企業横断比較）"
c.font = Font(bold=True, color="FFFFFF", size=13)
c.fill = fill("1A3A5C")
c.alignment = align("center", "center")
ws2.row_dimensions[1].height = 30

for ci, col_name in enumerate(SUM_COLS, 1):
    c = ws2.cell(row=2, column=ci, value=col_name)
    c.font = Font(bold=True, color="FFFFFF", size=10)
    c.fill = fill(C_COL_TITLE)
    c.alignment = align("center", "center")
    c.border = thin_border()
ws2.row_dimensions[2].height = 28

for i, w in enumerate(SUM_WIDTHS, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

for ri, (name, cat, shizuoka, star, strength) in enumerate(SUMMARY, 1):
    row_color = row_fill_color(cat)
    data = [str(ri), name, cat, shizuoka, star, strength]
    for ci, val in enumerate(data, 1):
        c = ws2.cell(row=ri + 2, column=ci, value=val)
        c.border = thin_border()
        c.alignment = align("left", "center")
        c.fill = fill(row_color)
        if ci == 1:
            c.font = Font(bold=True, size=10)
            c.alignment = align("center", "center")
        elif ci == 5:
            c.fill = fill(STAR_COLOR.get(star, C_STAR1))
            c.font = Font(bold=True, size=11,
                          color="FFFFFF" if star in ["★★★★★", "★★★★"] else "333333")
            c.alignment = align("center", "center")
        else:
            c.font = Font(size=9)
    ws2.row_dimensions[ri + 2].height = 28

# ---- シート3: 凡例・説明 ----
ws3 = wb.create_sheet("凡例・説明")

ws3["A1"] = "■ DX成熟度の定義"
ws3["A1"].font = Font(bold=True, size=12, color="FFFFFF")
ws3["A1"].fill = fill("1A3A5C")
ws3.merge_cells("A1:C1")
ws3.row_dimensions[1].height = 24

legend_dx = [
    ("DX成熟度", "基準", "代表企業"),
    ("★★★★★", "複数のAI施策・データ基盤・リテールメディア事業化が揃っている", "ウエルシア・スギ・マツキヨ"),
    ("★★★★",  "AI活用・データ基盤・デジタル施策が高水準で進んでいる",         "サンドラッグ・ドンキ・イオン"),
    ("★★★",   "基本的なアプリ・デジタル施策は整備済み",                        "ツルハHD・クリエイトSD・クスリのアオキ"),
    ("★★",    "デジタル施策は限定的、または情報が少ない",                       "ゲンキー・カワチ薬品"),
    ("★",     "情報なし・未着手、またはDX優先度が低い",                        "田子重・業務スーパー"),
]
for ri, (star, basis, example) in enumerate(legend_dx, 2):
    color = STAR_COLOR.get(star, C_COL_TITLE)
    is_header = ri == 2
    for ci, val in enumerate([star, basis, example], 1):
        c = ws3.cell(row=ri, column=ci, value=val)
        c.border = thin_border()
        c.alignment = align("left", "center")
        if is_header:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = fill(C_COL_TITLE)
        else:
            c.fill = fill(color)
            is_dark = star in ["★★★★★", "★★★★"]
            c.font = Font(bold=(ci == 1), size=11 if ci == 1 else 9,
                          color="FFFFFF" if (is_dark and ci == 1) else "333333")
    ws3.row_dimensions[ri].height = 28

ws3.cell(row=9, column=1, value="■ 分類（トピック分類）の定義")
ws3["A9"].font = Font(bold=True, size=12, color="FFFFFF")
ws3["A9"].fill = fill("1A3A5C")
ws3.merge_cells("A9:C9")
ws3.row_dimensions[9].height = 24

legend_class = [
    ("分類", "対象となる情報"),
    ("AI活用", "AI導入事例・PoC・生成AI・機械学習・自動化"),
    ("データ基盤", "CDP・DWH・データレイク・データ連携・MDM"),
    ("アプリ・デジタル", "公式アプリ・EC・デジタルサービス・サイネージ"),
    ("調剤DX", "処方箋送信・電子お薬手帳・服薬指導・調剤システム"),
    ("リテールメディア", "広告プラットフォーム・サイネージ広告・ファーストパーティデータ活用"),
    ("EX（従業員体験）", "従業員向けDXツール・業務効率化・ナレッジ管理"),
    ("その他", "上記に当てはまらないデジタル施策"),
]
for ri, row_data in enumerate(legend_class, 10):
    is_header = ri == 10
    for ci, val in enumerate(row_data, 1):
        c = ws3.cell(row=ri, column=ci, value=val)
        c.border = thin_border()
        c.alignment = align("left", "center")
        if is_header:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = fill(C_COL_TITLE)
        else:
            c.font = Font(bold=(ci == 1), size=9)
    ws3.row_dimensions[ri].height = 24

ws3.cell(row=19, column=1, value="■ 実装時期の表記ルール")
ws3["A19"].font = Font(bold=True, size=12, color="FFFFFF")
ws3["A19"].fill = fill("1A3A5C")
ws3.merge_cells("A19:C19")
ws3.row_dimensions[19].height = 24

legend_timing = [
    ("表記", "意味"),
    ("YYYY年MM月", "公表情報から確認できた具体的な導入・完了時期"),
    ("YYYY年頃", "推定または概算の時期"),
    ("YYYY年予定", "今後の導入予定が公表されているもの"),
    ("構築中 / 実証中 / 継続中", "現在進行中の取り組み"),
    ("検討中", "導入・実施が検討段階にあるもの"),
    ("不明", "公表情報から時期が確認できないもの"),
]
for ri, row_data in enumerate(legend_timing, 20):
    is_header = ri == 20
    for ci, val in enumerate(row_data, 1):
        c = ws3.cell(row=ri, column=ci, value=val)
        c.border = thin_border()
        c.alignment = align("left", "center")
        if is_header:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = fill(C_COL_TITLE)
        else:
            c.font = Font(bold=(ci == 1), size=9)
    ws3.row_dimensions[ri].height = 24

ws3.cell(row=28, column=1, value="■ ツルハHD比較の意味")
ws3["A28"].font = Font(bold=True, size=12, color="FFFFFF")
ws3["A28"].fill = fill("1A3A5C")
ws3.merge_cells("A28:C28")
ws3.row_dimensions[28].height = 24

legend_tsurha = [
    ("ツルハHD比較", "意味"),
    ("未着手", "ツルハHDが現時点で取り組んでいない施策。杏林堂にとって競合に遅れているリスク領域"),
    ("同等", "ツルハHDが同等レベルで取り組んでいる施策"),
    ("不明", "ツルハHDの取り組み状況が確認できていない"),
    ("-", "ツルハHD自身の施策（参考行）"),
]
for ri, row_data in enumerate(legend_tsurha, 29):
    is_header = ri == 29
    for ci, val in enumerate(row_data, 1):
        c = ws3.cell(row=ri, column=ci, value=val)
        c.border = thin_border()
        c.alignment = align("left", "center")
        if is_header:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = fill(C_COL_TITLE)
        elif ri == 30 and ci == 1:
            c.font = Font(bold=True, color="CC0000")
            c.fill = fill(C_MIZEN)
        else:
            c.font = Font(size=9)
    ws3.row_dimensions[ri].height = 28

for col, w in [("A", 22), ("B", 60), ("C", 32)]:
    ws3.column_dimensions[col].width = w

# ---- 保存 ----
OUTPUT = (
    r"C:\Users\takatoshi-saito\OneDrive\00personal\ClaudeCodeFolder"
    r"\kyorindo-cx\04_references\競合DXベンチマーク_20260423.xlsx"
)
wb.save(OUTPUT)
print(f"保存完了: {OUTPUT}")

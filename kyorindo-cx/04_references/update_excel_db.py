import sys, os
from datetime import date

sys.stdout.reconfigure(encoding='utf-8')

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    os.system('pip install openpyxl -q')
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'ツルハHD_IT記事データベース.xlsx')
COLUMNS = ['掲載日', 'タイトル', 'カテゴリ', '媒体名', 'URL', '概要', 'キーワード', '関連ベンダー/企業', 'MDファイル名', '追加日']

if not os.path.exists(EXCEL_PATH):
    wb = Workbook()
    ws = wb.active
    ws.title = '記事一覧'
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color='1F3864')
        cell.alignment = Alignment(horizontal='center')
    widths = [12, 45, 14, 20, 60, 60, 35, 30, 50, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    wb.save(EXCEL_PATH)
    print('新規作成:', EXCEL_PATH)

wb = load_workbook(EXCEL_PATH)
ws = wb.active

existing_urls = set()
existing_md_names = set()
url_col = COLUMNS.index('URL') + 1
md_col = COLUMNS.index('MDファイル名') + 1
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[url_col - 1]:
        existing_urls.add(str(row[url_col - 1]).strip())
    if row[md_col - 1]:
        existing_md_names.add(str(row[md_col - 1]).strip())

today_str = date.today().strftime('%Y/%m/%d')

all_articles = [
    {
        '掲載日': '2026/04/09',
        'タイトル': 'ツルハHD 中期経営計画（2027年2月期〜2029年2月期）策定：LIFE STOREビジョンとAIエージェント活用',
        'カテゴリ': '決算・経営',
        '媒体名': 'Biz/Zine / 流通ニュース',
        'URL': 'https://bizzine.jp/news/detail/12910',
        '概要': '売上高2.7兆円・営業利益1350億円を目標とした中計を発表。AIエージェント活用・LIFE STOREビジョン・3年1500店舗出店・PB15%を柱に。',
        'キーワード': '中期経営計画 LIFE STORE AIエージェント 2.7兆円 PB15% PHASE1',
        '関連ベンダー/企業': '',
        'MDファイル名': 'ツルハHD_中期経営計画_LIFESTOREビジョン_20260409.md',
    },
    {
        '掲載日': '2026/04/20',
        'タイトル': 'ツルハHD 2026年2月期決算：ウエルシアHDとの統合後初、次期2兆5550億円企業へ',
        'カテゴリ': '決算・経営',
        '媒体名': 'ドラッグストア＆調剤（HPC News）',
        'URL': 'https://www.hpc-news.co.jp/media/info/a2033',
        '概要': '統合後初決算・売上1兆4505億円。次期2兆5550億円予想・シナジー3カ年400億円・1000万時間削減目標・統合ヘルスケアデータベース活用を発表。',
        'キーワード': '決算 2026年2月期 統合後初 シナジー 1000万時間削減 統合ヘルスケアDB',
        '関連ベンダー/企業': '',
        'MDファイル名': 'ツルハHD_2026年2月期決算_統合後初_シナジー計画_20260420.md',
    },
    {
        '掲載日': '2026/04/17',
        'タイトル': 'ツルハHD DX戦略：顧客ID統一で共通アプリリリース＋ECサイト集約で顧客体験刷新',
        'カテゴリ': 'データ基盤',
        '媒体名': 'ネットショップ担当者フォーラム',
        'URL': 'https://netshop.impress.co.jp/n/2026/04/17/15940',
        '概要': '顧客ID統一・共通アプリ・EC集約・AIエージェント（3年後提供目標）・WAON POINTからの移行計画を詳説。',
        'キーワード': '顧客ID統一 共通アプリ EC集約 AIエージェント WAON POINT DX戦略',
        '関連ベンダー/企業': '',
        'MDファイル名': 'ツルハHD_DX戦略_共通アプリ顧客ID統一_20260417.md',
    },
    {
        '掲載日': '2026/03/24',
        'タイトル': 'ツルハHD 生成AI社内ナレッジ検索：NEC・ProofXと共同で1000店舗以上に導入',
        'カテゴリ': 'AI活用',
        '媒体名': 'PR TIMES',
        'URL': 'https://prtimes.jp/main/html/rd/p/000001274.000078149.html',
        '概要': 'NEC・ProofXと共同で生成AI活用の社内ナレッジ検索システムを1000店舗以上に展開。店舗スタッフの業務効率化を図る。',
        'キーワード': '生成AI ナレッジ検索 NEC ProofX 1000店舗 店舗DX',
        '関連ベンダー/企業': 'NEC / ProofX',
        'MDファイル名': 'ツルハHD_生成AI社内ナレッジ検索_NEC_ProofX_20260324.md',
    },
    {
        '掲載日': '2026/03/02',
        'タイトル': 'ウエルシア薬局 業務端末をAndroidスマホへ全面リプレイス：ソフトバンクと協業',
        'カテゴリ': '店舗DX',
        '媒体名': 'ダイヤモンド・チェーンストアオンライン',
        'URL': 'https://diamond-rm.net/promotion/pr/534995/',
        '概要': 'ウエルシア薬局がAndroidスマートフォンへ業務端末を全面リプレイス。ソフトバンクとの協業で店舗オペレーション効率化を推進。',
        'キーワード': 'Android リプレイス ソフトバンク 業務端末 店舗DX ウエルシア',
        '関連ベンダー/企業': 'ソフトバンク',
        'MDファイル名': 'ツルハHD_ウエルシア業務端末Androidリプレイス_ソフトバンク_20260302.md',
    },
    {
        '掲載日': '2025/12/02',
        'タイトル': 'ツルハHD・ウエルシア データクリーンルーム統合：国内最大級リテールメディア基盤を構築',
        'カテゴリ': 'リテールメディア',
        '媒体名': 'PR TIMES（アドインテ）',
        'URL': 'https://prtimes.jp/main/html/rd/p/000000212.000007452.html',
        '概要': 'ツルハHDとウエルシアのDCRを統合し国内最大級のリテールメディア基盤を構築。メーカー向け広告配信・効果計測を統合環境で提供。',
        'キーワード': 'データクリーンルーム DCR リテールメディア アドインテ 統合',
        '関連ベンダー/企業': 'アドインテ（AdInte）',
        'MDファイル名': 'ツルハHD_データクリーンルーム_リテールメディア_20251202.md',
    },
    {
        '掲載日': '2025/07/08',
        'タイトル': 'ツルハHD SmartDB導入：店舗マスタ・契約管理台帳を全社展開でIT基盤DXを推進',
        'カテゴリ': 'データ基盤',
        '媒体名': 'ドリーム・アーツ（プレスリリース）',
        'URL': 'https://www.dreamarts.co.jp/news/press-release/pr250708/',
        '概要': 'SmartDB(R)で店舗マスタ・契約管理台帳を全社展開。200項目超の店舗情報を一元管理し、新リース会計基準にも対応。',
        'キーワード': 'SmartDB 店舗マスタ 契約管理 ドリームアーツ IT基盤 内部統制',
        '関連ベンダー/企業': 'ドリーム・アーツ',
        'MDファイル名': 'ツルハHD_SmartDB導入_IT基盤DX_20250708.md',
    },
    {
        '掲載日': '2025/02/14',
        'タイトル': 'ウエルシア薬局 SalesSensor導入：True DataのAIで新規出店売上予測を自動化',
        'カテゴリ': 'AI活用',
        '媒体名': 'PR TIMES（True Data）',
        'URL': 'https://prtimes.jp/main/html/rd/p/000000132.000039871.html',
        '概要': 'ウエルシア薬局がTrue DataのAI「SalesSensor」を2025年度導入。商圏・人流・競合データを統合し物販＋調剤の両軸で売上予測を自動化。',
        'キーワード': 'SalesSensor True Data 新規出店 売上予測 AI 商圏分析 調剤',
        '関連ベンダー/企業': 'True Data / DATAFLUCT',
        'MDファイル名': 'ツルハHD_ウエルシアSalesSensor新規出店売上予測AI_TrueData_20250214.md',
    },
    {
        '掲載日': '2024/12/19',
        'タイトル': 'ウエルシア薬局 カスタマーAI導入：NELで1億件の顧客の声を分析しPB商品開発を効率化',
        'カテゴリ': 'AI活用',
        '媒体名': 'PR TIMES（NEL）',
        'URL': 'https://prtimes.jp/main/html/rd/p/000000123.000087636.html',
        '概要': 'NELの「カスタマーAI」で1億件超の顧客意見を分析。CS業務90%削減・PB商品企画のAI起点化を約6か月で実現。',
        'キーワード': 'カスタマーAI NEL VoC分析 PB商品開発 CS削減90% ウエルシア',
        '関連ベンダー/企業': 'NEL',
        'MDファイル名': 'ツルハHD_ウエルシアカスタマーAI_NEL_PB商品開発_20241219.md',
    },
    {
        '掲載日': '2024/12/04',
        'タイトル': 'ウエルシア薬局 AI販促ソリューション導入：True DataのAIで1to1クーポン配信',
        'カテゴリ': 'AI活用',
        '媒体名': 'True Data（プレスリリース）',
        'URL': 'https://www.truedata.co.jp/release20241204/',
        '概要': 'ウエルシアがTrue DataのAI販促ソリューションを2025年度導入。購買データをもとに個人最適化クーポンを配信。',
        'キーワード': 'AI販促 True Data 1to1 クーポン 購買データ ウエルシア',
        '関連ベンダー/企業': 'True Data',
        'MDファイル名': 'ツルハHD_ウエルシアAI販促_TrueData_20241204.md',
    },
    {
        '掲載日': '2024/05/16',
        'タイトル': 'ツルハドラッグ AI薬歴入力支援：ChatGPT×音声認識AIで薬歴作成時間を最大68%削減',
        'カテゴリ': '調剤・医療',
        '媒体名': 'PR TIMES（WEMEX）',
        'URL': 'https://prtimes.jp/main/html/rd/p/000000073.000107062.html',
        '概要': 'WEMEXがChatGPT×音声認識AIで薬歴入力支援システムを開発。ツルハドラッグ実店舗でトライアル開始、薬歴作成時間を最大68%削減。',
        'キーワード': 'AI薬歴 WEMEX ChatGPT 音声認識 薬歴作成68%削減 調剤',
        '関連ベンダー/企業': 'WEMEX',
        'MDファイル名': 'ツルハHD_AI薬歴入力支援_WEMEX_ChatGPT_20240516.md',
    },
    {
        '掲載日': '2023/11/20',
        'タイトル': 'ツルハグループ リテールメディア広告事業：ツルハグループAdsで億単位の広告事業を確立',
        'カテゴリ': 'リテールメディア',
        '媒体名': '日経ビジネス',
        'URL': 'https://business.nikkei.com/atcl/gen/19/00592/111700003/',
        '概要': 'ツルハグループAds（2020年8月開始）で年間1200万人の購買データを活用した億単位の広告事業を構築。',
        'キーワード': 'リテールメディア ツルハグループAds 購買データ 広告事業 1200万人',
        '関連ベンダー/企業': '',
        'MDファイル名': 'ツルハHD_リテールメディア広告事業_ツルハグループAds_20231120.md',
    },
]

added = 0
for article in all_articles:
    url = article.get('URL', '').strip()
    md_name = article.get('MDファイル名', '').strip()
    if url in existing_urls or md_name in existing_md_names:
        continue
    row_data = [
        article.get('掲載日', ''),
        article.get('タイトル', ''),
        article.get('カテゴリ', ''),
        article.get('媒体名', ''),
        url,
        article.get('概要', ''),
        article.get('キーワード', ''),
        article.get('関連ベンダー/企業', ''),
        md_name,
        today_str,
    ]
    ws.append(row_data)
    existing_urls.add(url)
    existing_md_names.add(md_name)
    added += 1

wb.save(EXCEL_PATH)
print(f'Excel更新完了: {added}件追加')
print(f'保存先: {EXCEL_PATH}')

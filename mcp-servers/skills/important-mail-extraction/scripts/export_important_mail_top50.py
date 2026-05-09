import argparse
import json
import re
import sqlite3
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[3]
CYBERMAIL_DIR = ROOT / "cybermail-imap"
DEFAULT_DB = CYBERMAIL_DIR / "kb" / "emails.db"
DEFAULT_RULES = CYBERMAIL_DIR / "sender_rules.xlsx"
DEFAULT_OUTPUT_DIR = CYBERMAIL_DIR / "outputs"


KEYWORDS = [
    "依頼",
    "確認",
    "対応",
    "相談",
    "期限",
    "承認",
    "回答",
    "返信",
    "障害",
    "エラー",
    "不具合",
    "復旧",
    "請求",
    "支払",
    "契約",
    "見積",
    "会議",
    "資料",
    "日程",
    "調整",
    "故障",
    "受付",
    "連絡",
    "至急",
    "OPEN21 Workflow",
    "滞留通知",
    "ワークフロー",
    "申請",
    "催促",
    "完了",
    "宿題",
    "情報セキュリティ",
    "体制図",
    "運用開始",
    "部内へ共有",
    "準備完了",
    "精算",
    "取り消し",
    "取消",
    "精算戻し",
]

STRONG_KEYWORDS = [
    "故障",
    "障害",
    "エラー",
    "不具合",
    "復旧",
    "滞留通知",
    "承認",
    "至急",
    "請求",
    "支払",
    "契約",
    "見積",
    "修理",
    "回答",
    "依頼",
    "宿題",
    "期日",
    "情報セキュリティ",
    "セキュリティ管理委員会",
    "体制図",
    "執行役員会議",
    "運用開始",
    "部内へ共有",
    "準備完了",
    "精算",
    "取り消し",
    "取消",
    "精算戻し",
    "機能についてはいかがでしょうか",
]

DEADLINE_PATTERNS = [
    r"\d{1,2}/\d{1,2}\s*[（(]?[月火水木金土日]?[）)]?",
    r"\d{1,2}月\d{1,2}日",
    r"\d{4}[-/]\d{1,2}[-/]\d{1,2}",
    r"本日",
    r"明日",
    r"今週中",
    r"来週中",
    r"月末",
]

NOISE_SUBJECTS = [
    "自動配信",
    "リマインドメール",
    "自動作成連絡",
    "粗利金額推移",
    "日次生産性",
    "ログイン通知",
    "メルマガ",
    "広告",
    "日報作成完了",
    "勤怠管理システム",
]


def extract_emails(text):
    return [
        item.lower()
        for item in re.findall(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+", text or "")
    ]


def email_domain(address):
    return address.split("@")[-1].lower() if "@" in address else ""


def unique(items):
    return list(dict.fromkeys([item for item in items if item]))


def load_rules(rules_path):
    workbook = openpyxl.load_workbook(rules_path, read_only=True, data_only=True)
    personal = set()
    domains = {}

    for sheet_name in ["PersonalAddressRules", "DomainRules"]:
        worksheet = workbook[sheet_name]
        headers = [
            cell
            for cell in next(
                worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
            )
        ]
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            values = dict(zip(headers, row))
            if not values.get("enabled"):
                continue
            pattern = str(values.get("pattern") or "").lower().strip()
            match_type = str(values.get("match_type") or "")
            sender_type = str(values.get("sender_type") or "").strip()
            if not pattern:
                continue
            if sheet_name == "PersonalAddressRules" and (
                match_type == "exact_from_addr" or "@" in pattern
            ):
                personal.add(pattern)
            if sheet_name == "DomainRules" and sender_type in (
                "internal",
                "important_vendor",
            ):
                domains[pattern] = sender_type

    return personal, domains


def classify(subject, body, sender_type):
    text = f"{subject} {body[:1200]}"
    if any(word in text for word in ["障害", "エラー", "不具合", "復旧", "故障", "修理"]):
        return "障害・メンテ"
    if any(word in text for word in ["請求", "支払", "契約", "見積"]):
        return "請求・契約"
    if any(word in text for word in ["Workflow", "ワークフロー", "承認", "滞留通知", "申請"]):
        return "対応必要"
    if sender_type == "internal":
        return "社内重要"
    if sender_type == "important_vendor":
        return "重要ベンダー"
    if any(word in text for word in ["日程", "会議", "資料", "返信", "回答", "確認", "対応", "依頼", "相談"]):
        return "対応必要"
    return "重要情報"


def recommended_action(classification):
    actions = {
        "障害・メンテ": "受付内容と対応指示・復旧状況を確認",
        "請求・契約": "金額・期限・承認要否を確認",
        "対応必要": "未回答・未承認・期限の有無を確認して対応",
        "重要ベンダー": "取引先対応要否を確認",
        "社内重要": "社内共有または対応要否を確認",
    }
    return actions.get(classification, "内容確認")


def find_deadline_hits(text):
    if not any(word in text for word in ["期日", "期限", "まで", "締切", "締め切り"]):
        return []
    hits = []
    for pattern in DEADLINE_PATTERNS:
        hits.extend(re.findall(pattern, text))
    return unique(hits)[:5]


def score_mail(row, personal_rules, domain_rules):
    subject = row["subject"] or ""
    body = row["body_text"] or ""
    from_to_text = f"{row['from_addr'] or ''} {row['to_addr'] or ''}"
    addresses = set(extract_emails(from_to_text))
    body_addresses = set(extract_emails(body[:3000]))

    rule_hits = []
    hit_types = []

    for address in addresses:
        if address in personal_rules:
            rule_hits.append(f"personal:{address}")
            hit_types.append("personal")
        domain = email_domain(address)
        if domain in domain_rules:
            sender_type = domain_rules[domain]
            rule_hits.append(f"{sender_type}:{domain}")
            hit_types.append(sender_type)

    for address in body_addresses:
        domain = email_domain(address)
        if address in personal_rules:
            rule_hits.append(f"body/Cc候補 personal:{address}")
        elif domain in domain_rules:
            rule_hits.append(f"body/Cc候補 {domain_rules[domain]}:{domain}")

    text = f"{subject} {body[:1800]}"
    keyword_hits = [word for word in KEYWORDS if word in text]
    strong_hits = [word for word in STRONG_KEYWORDS if word in text]
    deadline_hits = find_deadline_hits(text)
    is_noise = any(word in subject for word in NOISE_SUBJECTS)
    has_rule = bool(rule_hits)

    if not has_rule and not (keyword_hits and (strong_hits or deadline_hits) and not is_noise):
        return None

    sender_type = row["sender_type"] or ""
    if "important_vendor" in hit_types or any("important_vendor" in hit for hit in rule_hits):
        sender_type = "important_vendor"
    elif "internal" in hit_types or any("internal" in hit for hit in rule_hits):
        sender_type = "internal"

    extraction_reason = "sender_rules"
    if has_rule and (strong_hits or deadline_hits):
        extraction_reason = "複合"
    elif not has_rule:
        extraction_reason = "AI判定"

    score = (
        (100 if has_rule else 0)
        + len(keyword_hits) * 4
        + len(strong_hits) * 8
        + len(deadline_hits) * 10
        + (20 if any(word in text for word in ["故障", "障害", "滞留通知", "承認", "至急", "修理"]) else 0)
        + (24 if any(word in text for word in ["宿題", "情報セキュリティ", "体制図", "セキュリティ管理委員会"]) else 0)
        + (20 if any(word in text for word in ["運用開始", "部内へ共有", "準備完了"]) else 0)
        + (20 if any(word in text for word in ["精算", "取り消し", "取消", "精算戻し"]) else 0)
        - (50 if is_noise and not has_rule else 0)
    )
    classification = classify(subject, body, sender_type)
    important_reason = " / ".join(unique(rule_hits + strong_hits + [f"期限候補:{hit}" for hit in deadline_hits])[:10])

    return {
        "score": score,
        "date": row["date"],
        "from_addr": row["from_addr"],
        "to_addr": row["to_addr"],
        "to_cc_important_match": " / ".join(unique(rule_hits)[:8]),
        "subject": subject,
        "classification": classification,
        "extraction_reason": extraction_reason,
        "important_reason": important_reason or "内容キーワード一致",
        "recommended_action": recommended_action(classification),
        "body_preview": re.sub(r"\s+", " ", body)[:300],
        "email_id": row["id"],
    }


def collect_candidates(db_path, rules_path, since):
    personal_rules, domain_rules = load_rules(rules_path)
    connection = sqlite3.connect(db_path)
    connection.row_factory = sqlite3.Row
    rows = connection.execute(
        """
        select id, date, folder, from_addr, to_addr, subject, body_text,
               sender_type, classification, attachments
        from emails
        where date >= ?
        order by date desc
        """,
        (since,),
    ).fetchall()

    candidates = []
    for row in rows:
        scored = score_mail(row, personal_rules, domain_rules)
        if scored:
            candidates.append(scored)

    candidates.sort(key=lambda item: (item["score"], item["date"] or ""), reverse=True)
    return candidates, len(rows)


def write_excel(candidates, output_path):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "important_mail"

    headers = [
        "rank",
        "score",
        "date",
        "from_addr",
        "to_addr",
        "to_cc_important_match",
        "subject",
        "classification",
        "extraction_reason",
        "important_reason",
        "recommended_action",
        "body_preview",
        "email_id",
    ]
    worksheet.append(headers)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for index, candidate in enumerate(candidates, start=1):
        worksheet.append([index] + [candidate.get(header, "") for header in headers[1:]])

    widths = {
        "A": 8,
        "B": 8,
        "C": 18,
        "D": 34,
        "E": 34,
        "F": 42,
        "G": 54,
        "H": 16,
        "I": 16,
        "J": 48,
        "K": 34,
        "L": 80,
        "M": 10,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{worksheet.max_row}"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--since", required=True, help="検索開始日。例: 2026-05-06")
    parser.add_argument("--limit", type=int, default=0, help="0または未指定で候補すべてを出力")
    parser.add_argument("--db", default=str(DEFAULT_DB))
    parser.add_argument("--rules", default=str(DEFAULT_RULES))
    parser.add_argument("--out")
    return parser.parse_args()


def main():
    args = parse_args()
    db_path = Path(args.db)
    rules_path = Path(args.rules)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = (
        Path(args.out)
        if args.out
        else DEFAULT_OUTPUT_DIR / f"important_mail_all_{timestamp}.xlsx"
    )

    candidates, target_count = collect_candidates(db_path, rules_path, args.since)
    selected = candidates if args.limit <= 0 else candidates[: args.limit]
    write_excel(selected, output_path)

    print(
        json.dumps(
            {
                "output": str(output_path),
                "since": args.since,
                "target_count": target_count,
                "candidate_count": len(candidates),
                "exported_count": len(selected),
                "cc_note": "DBにcc_addr列がない場合、本文内ヘッダ相当テキストからCc候補を判定",
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
